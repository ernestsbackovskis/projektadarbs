import imaplib, email, os, smtplib 
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import datetime 
from openpyxl import Workbook, load_workbook

todays = datetime.date.today()
user="ernestiks131@gmail.com"
passw="hrxe eezk zlln rynu"
imap_url ="imap.gmail.com"
key = 'FROM'
value = 'ernestiks131@gmail.com'
key2="SUBJECT"
value2=("Inventerizacija"+str(todays))
attach_dir= 'D:/Python/Lietojumprogrammaturasariks/projekta_darbs/Lejup/'

def get_attach(msg):
    for part in msg.walk():
        if part.get_content_maintype()=="multipart":
            continue
        if part.get('Content-Disposition') is None:
            continue
        fileName = part.get_filename()

        if bool(fileName):
            filePath = os.path.join(attach_dir,fileName)
            with open(filePath,'wb') as f:
                f.write(part.get_payload(decode=True))


def get_body(msg):
    if msg.is_multipart():
        return get_body(msg.get_payload(0))
    else:
        return msg.get_payload(None,True)

con=imaplib.IMAP4_SSL(imap_url)
con.login(user, passw)
con.select("INBOX")
resp, items = con.search(None, key, value, key2, value2)
items = items[0].split()
for emailid in items:
    result, data = con.fetch(emailid,"(RFC822)")
    raw = email.message_from_bytes(data[0][1])
    for part in raw.walk():
        if part.get_content_maintype()=="multipart":
            continue
        if part.get('Content-Disposition') is None:
            continue
        fileName = part.get_filename()

        if bool(fileName):
            filePath = os.path.join(attach_dir,fileName)
            with open(filePath,'wb') as f:
                f.write(part.get_payload(decode=True))
    
wb=load_workbook("PC.xlsx")
ws=wb.active
wb1=load_workbook(filePath)
ws1=wb1.active
wb2=Workbook()
ws2=wb2.active
max_row=ws.max_row
max_row1=ws1.max_row
name=[]
lastName=[]
serial=[]
date12=[]
email=[]
today = datetime.date.today()-datetime.timedelta(days=4*365)

for i in range(1,max_row):
    date=(ws['C'+str(i)].value)
    if str(today)>=date:
        name.append(ws['D'+str(i)].value)
        lastName.append(ws['E'+str(i)].value)
        serial.append(ws['B'+str(i)].value)
        date12.append(date)
(ws2['A'+str(1)].value)="Name"
(ws2['B'+str(1)].value)="Last Name"
(ws2['C'+str(1)].value)="Serial Number"
(ws2['D'+str(1)].value)="Purchese date"
(ws2['E'+str(1)].value)="Worker Email"
for i in range(0,len(name)):
    for j in range(1, max_row1):
        wornam=(ws1['B'+str(j)].value)
        worlasn=(ws1['C'+str(j)].value)
        if name[i]==wornam and lastName[i]==worlasn:
            (ws2['A'+str(i+2)].value)=name[i]
            (ws2['B'+str(i+2)].value)=lastName[i]
            (ws2['C'+str(i+2)].value)=serial[i]
            (ws2['D'+str(i+2)].value)=date12[i]
            (ws2['E'+str(i+2)].value)=(ws1['D'+str(j)].value)
sfil=("Utilizacija"+str(todays)+".xlsx")
wb2.save(sfil)
wb.close()
wb1.close()
wb2.close()

gmail_server= "smtp.gmail.com"
gmail_port= 587

msg=MIMEMultipart()
msg["From"]="ernestiks131@gmail.com"
msg["To"]="ernestiks131@gmail.com"
msg["Subject"]=("Inventerizacijasatskaite"+str(todays))

body="Te ir atskaite par datoriem, kas ir jāutilizē. \n Ar cieņu, \n Ernests Bačkovskis"
msg.attach(MIMEText(body,'plain'))

attachment = open(sfil,'rb')
part=MIMEBase('application','octet-stream')
part.set_payload((attachment).read())
encoders.encode_base64(part)
part.add_header('Content-Disposition',"attachment; filename="+sfil)

msg.attach(part)
text=msg.as_string()
my_server = smtplib.SMTP(gmail_server, gmail_port)
my_server.starttls()
my_server.login(user,passw)



my_server.sendmail("ernestiks131@gmail.com","ernestiks131@gmail.com",text)
my_server.quit()